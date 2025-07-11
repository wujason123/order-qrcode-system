#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
订单查询Flask后端服务
"""

from flask import Flask, request, jsonify, render_template, send_from_directory, send_file, redirect, url_for, flash, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
import sqlite3
import os
from datetime import datetime
from excel_processor import OrderProcessor
import time
import pandas as pd

app = Flask(__name__)
CORS(app)  # 允许跨域请求
app.secret_key = 'order-qrcode-system-secret-key-2024'  # 用于session和flash消息

# 用户配置 - 可以从环境变量或配置文件读取
USERS = {
    'admin': generate_password_hash('admin123'),  # 默认管理员账户
    'manager': generate_password_hash('manager123')  # 默认管理员账户
}

# 配置
DB_FILE = "orders.db"
QR_DIR = "qrcodes"
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 确保上传目录存在
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 认证相关函数
def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            if request.is_json:
                return jsonify({'error': '需要登录访问', 'redirect': '/login'}), 401
            flash('请先登录才能访问此页面', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def authenticate_user(username, password):
    """验证用户名和密码"""
    if username in USERS:
        return check_password_hash(USERS[username], password)
    return False

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db_connection():
    """获取数据库连接"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row  # 使返回结果可以像字典一样访问
    return conn

# 认证相关路由
@app.route('/login', methods=['GET', 'POST'])
def login():
    """用户登录"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username and password:
            if authenticate_user(username, password):
                session['logged_in'] = True
                session['username'] = username
                flash(f'欢迎回来，{username}！', 'success')
                
                # 获取登录前要访问的页面
                next_page = request.args.get('next')
                if next_page:
                    return redirect(next_page)
                return redirect(url_for('index'))
            else:
                flash('用户名或密码错误', 'error')
        else:
            flash('请输入用户名和密码', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """用户登出"""
    username = session.get('username', '用户')
    session.clear()
    flash(f'再见，{username}！', 'info')
    return redirect(url_for('login'))

@app.route('/')
def index():
    """主页 - 重定向到登录页面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/public')
def public_query():
    """公共查询页面 - 无需登录，只能查询订单"""
    return render_template('public.html')

@app.route('/print')
def print_page():
    """打印预览页面"""
    return render_template('print.html')

@app.route('/order')
def get_order():
    """获取订单信息API"""
    order_id = request.args.get('order_id')
    
    if not order_id:
        return jsonify({
            'error': '缺少订单号参数',
            'message': '请提供order_id参数'
        }), 400
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 查询订单信息
        cursor.execute('''
            SELECT order_id, customer_name, order_date, amount, product_details, created_at
            FROM orders 
            WHERE order_id = ?
        ''', (order_id,))
        
        order = cursor.fetchone()
        conn.close()
        
        if order:
            # 转换为字典格式
            order_data = {
                'order_id': order['order_id'],
                'customer_name': order['customer_name'],
                'order_date': order['order_date'],
                'amount': order['amount'],
                'product_details': order['product_details'],
                'created_at': order['created_at']
            }
            
            return jsonify({
                'success': True,
                'data': order_data
            })
        else:
            return jsonify({
                'error': '订单不存在',
                'message': f'未找到订单号为 {order_id} 的订单'
            }), 404
            
    except Exception as e:
        return jsonify({
            'error': '服务器错误',
            'message': str(e)
        }), 500

@app.route('/orders')
@login_required
def list_orders():
    """获取所有订单列表（管理用）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT order_id, customer_name, order_date, amount, product_details
            FROM orders 
            ORDER BY order_date DESC
        ''')
        
        orders = cursor.fetchall()
        conn.close()
        
        orders_list = []
        for order in orders:
            orders_list.append({
                'order_id': order['order_id'],
                'customer_name': order['customer_name'],
                'order_date': order['order_date'],
                'amount': order['amount'],
                'product_details': order['product_details']
            })
        
        return jsonify({
            'success': True,
            'data': orders_list,
            'count': len(orders_list)
        })
        
    except Exception as e:
        return jsonify({
            'error': '服务器错误',
            'message': str(e)
        }), 500

@app.route('/qrcode/<order_id>')
def get_qrcode(order_id):
    """获取订单二维码图片"""
    try:
        filename = f"order_{order_id}.png"
        return send_from_directory(QR_DIR, filename)
    except Exception as e:
        return jsonify({
            'error': '二维码不存在',
            'message': f'订单 {order_id} 的二维码不存在'
        }), 404

@app.route('/health')
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """上传Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件被上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_with_timestamp = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_with_timestamp)
            
            file.save(filepath)
            
            # 处理上传的文件
            processor = OrderProcessor(excel_file=filepath)
            processor.init_database()
            result = processor.process_excel_data()
            
            if result['success']:
                # 生成二维码
                qr_result = processor.generate_qr_codes()
                if qr_result['success']:
                    return jsonify({
                        'success': True,
                        'message': f'成功处理 {result["count"]} 条订单数据，生成 {qr_result["count"]} 个二维码',
                        'orders_count': result['count'],
                        'qr_count': qr_result['count']
                    })
                else:
                    return jsonify({
                        'success': False,
                        'error': f'数据导入成功但二维码生成失败: {qr_result["error"]}'
                    }), 500
            else:
                # 返回详细的错误信息，包括重复订单号
                error_response = {
                    'success': False,
                    'error': result["error"]
                }
                
                # 如果有重复订单号信息，也一并返回
                if 'duplicates' in result:
                    error_response['duplicates'] = result['duplicates']
                    error_response['duplicate_type'] = result.get('type', 'excel_duplicate')
                
                return jsonify(error_response), 400
        else:
            return jsonify({'error': '文件格式不支持，请上传.xlsx或.xls文件'}), 400
            
    except Exception as e:
        return jsonify({'error': f'上传处理失败: {str(e)}'}), 500

@app.route('/upload_purchase', methods=['POST'])
@login_required
def upload_purchase_file():
    """上传采购订单Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件被上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_with_timestamp = f"purchase_{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_with_timestamp)
            
            file.save(filepath)
            
            # 处理采购订单文件
            processor = OrderProcessor()
            processor.init_database()
            result = processor.process_purchase_orders(filepath)
            
            if result['success']:
                return jsonify({
                    'success': True,
                    'message': f'成功处理 {result["count"]} 条采购订单，已自动计算订单盈亏状态',
                    'purchase_count': result['count']
                })
            else:
                return jsonify({
                    'success': False,
                    'error': result["error"]
                }), 400
        else:
            return jsonify({'error': '文件格式不支持，请上传.xlsx或.xls文件'}), 400
            
    except Exception as e:
        return jsonify({'error': f'采购订单上传处理失败: {str(e)}'}), 500

@app.route('/download/purchase_template')
@login_required
def download_purchase_template():
    """下载采购订单Excel模板"""
    try:
        processor = OrderProcessor()
        template_file = processor.create_purchase_sample_excel("purchase_template.xlsx")
        
        return send_file(
            template_file,
            as_attachment=True,
            download_name='采购订单模板.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'模板下载失败: {str(e)}'}), 500

@app.route('/download/bom_template')
@login_required
def download_bom_template():
    """下载BOM物料清单Excel模板"""
    try:
        processor = OrderProcessor()
        template_file = processor.create_bom_sample_excel("bom_template.xlsx")
        
        return send_file(
            template_file,
            as_attachment=True,
            download_name='BOM物料清单模板.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'BOM模板下载失败: {str(e)}'}), 500

@app.route('/upload_bom', methods=['POST'])
@login_required
def upload_bom_file():
    """上传BOM物料清单Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件被上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_with_timestamp = f"bom_{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_with_timestamp)
            
            file.save(filepath)
            
            # 处理BOM文件
            processor = OrderProcessor()
            processor.init_database()
            result = processor.process_bom_data(filepath)
            
            if result['success']:
                return jsonify({
                    'success': True,
                    'message': f'成功处理 {result["count"]} 条BOM记录',
                    'bom_count': result['count']
                })
            else:
                return jsonify({
                    'success': False,
                    'error': result["error"]
                }), 400
        else:
            return jsonify({'error': '文件格式不支持，请上传.xlsx或.xls文件'}), 400
            
    except Exception as e:
        return jsonify({'error': f'BOM文件上传处理失败: {str(e)}'}), 500

@app.route('/api/orders')
@login_required
def get_all_orders():
    """获取所有订单列表API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM orders ORDER BY order_date DESC')
        orders = cursor.fetchall()
        conn.close()
        
        orders_list = []
        for order in orders:
            orders_list.append({
                'order_id': order['order_id'],
                'customer_name': order['customer_name'],
                'order_date': order['order_date'],
                'amount': order['amount'],
                'product_details': order['product_details']
            })
        
        return jsonify({
            'success': True,
            'orders': orders_list,
            'count': len(orders_list)
        })
        
    except Exception as e:
        return jsonify({'error': f'获取订单列表失败: {str(e)}'}), 500

@app.route('/api/qrcodes')
def get_qr_codes():
    """获取二维码文件列表API"""
    try:
        if os.path.exists(QR_DIR):
            qr_files = [f for f in os.listdir(QR_DIR) if f.endswith('.png')]
            qr_list = []
            for qr_file in qr_files:
                order_id = qr_file.replace('order_', '').replace('.png', '')
                qr_list.append({
                    'order_id': order_id,
                    'filename': qr_file,
                    'url': f'/qrcode/{order_id}'
                })
            
            return jsonify({
                'success': True,
                'qrcodes': qr_list,
                'count': len(qr_list)
            })
        else:
            return jsonify({
                'success': True,
                'qrcodes': [],
                'count': 0
            })
            
    except Exception as e:
        return jsonify({'error': f'获取二维码列表失败: {str(e)}'}), 500

@app.route('/api/duplicates', methods=['DELETE'])
@login_required
def delete_duplicate_orders():
    """删除指定的重复订单"""
    try:
        data = request.get_json()
        if not data or 'order_ids' not in data:
            return jsonify({'error': '缺少order_ids参数'}), 400
        
        order_ids = data['order_ids']
        if not isinstance(order_ids, list) or len(order_ids) == 0:
            return jsonify({'error': 'order_ids必须是非空数组'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 删除指定的订单
        placeholders = ','.join(['?' for _ in order_ids])
        cursor.execute(f'DELETE FROM orders WHERE order_id IN ({placeholders})', order_ids)
        deleted_count = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        # 同时删除对应的二维码文件
        for order_id in order_ids:
            qr_file = f"{QR_DIR}/order_{order_id}.png"
            if os.path.exists(qr_file):
                os.remove(qr_file)
        
        return jsonify({
            'success': True,
            'message': f'成功删除 {deleted_count} 个重复订单',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        return jsonify({'error': f'删除订单失败: {str(e)}'}), 500

@app.route('/download/template')
def download_template():
    """下载Excel模板文件"""
    try:
        from excel_processor import OrderProcessor
        
        # 创建新格式的销售订单模板
        processor = OrderProcessor()
        template_file = processor.create_sales_order_sample_excel("template_orders.xlsx")
        
        return send_from_directory('.', template_file, as_attachment=True, 
                                 download_name='销售订单模板（支持盈亏计算）.xlsx')
        
    except Exception as e:
        return jsonify({'error': f'下载模板失败: {str(e)}'}), 500

@app.route('/export/qrcodes')
@login_required
def export_qrcodes_excel():
    """导出带二维码的Excel文件用于打印"""
    try:
        from excel_processor import OrderProcessor
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import io
        import tempfile
        
        # 尝试导入Image类，提供兼容性处理
        try:
            from openpyxl.drawing.image import Image as OpenpyxlImage
        except ImportError:
            try:
                from openpyxl.drawing import Image as OpenpyxlImage
            except ImportError:
                OpenpyxlImage = None
        
        # 获取所有订单数据
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT order_id, customer_name, order_date, amount, product_details
            FROM orders 
            ORDER BY order_date DESC
        ''')
        orders = cursor.fetchall()
        conn.close()
        
        if not orders:
            return jsonify({'error': '没有订单数据可导出'}), 404
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "订单二维码打印表"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 订单号
        ws.column_dimensions['B'].width = 12  # 客户姓名
        ws.column_dimensions['C'].width = 12  # 订单日期
        ws.column_dimensions['D'].width = 12  # 金额
        ws.column_dimensions['E'].width = 30  # 产品详情
        ws.column_dimensions['F'].width = 20  # 二维码
        
        # 设置标题行
        headers = ['订单号', '客户姓名', '订单日期', '金额', '产品详情', '二维码']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置行高
        ws.row_dimensions[1].height = 30
        
        # 填充数据和插入二维码
        current_row = 2
        for order in orders:
            order_id, customer_name, order_date, amount, product_details = order
            
            # 填充文本数据
            ws.cell(row=current_row, column=1, value=order_id)
            ws.cell(row=current_row, column=2, value=customer_name)
            ws.cell(row=current_row, column=3, value=order_date)
            ws.cell(row=current_row, column=4, value=f'¥{amount:.2f}')
            ws.cell(row=current_row, column=5, value=product_details)
            
            # 插入二维码图片
            qr_file = f"{QR_DIR}/order_{order_id}.png"
            if OpenpyxlImage is None:
                # 如果无法导入Image类，显示文本提示
                ws.cell(row=current_row, column=6, value=f'二维码: {order_id}')
            elif os.path.exists(qr_file):
                try:
                    img = OpenpyxlImage(qr_file)
                    # 调整图片大小（适合打印）
                    img.width = 80
                    img.height = 80
                    # 定位到F列
                    img.anchor = f'F{current_row}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"插入二维码图片失败: {e}")
                    ws.cell(row=current_row, column=6, value=f'二维码: {order_id}')
            else:
                ws.cell(row=current_row, column=6, value=f'二维码: {order_id}')
            
            # 设置行高（如果有图片则高一些，否则正常高度）
            if OpenpyxlImage is not None and os.path.exists(qr_file):
                ws.row_dimensions[current_row].height = 65
            else:
                ws.row_dimensions[current_row].height = 30
            
            # 设置单元格对齐
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            current_row += 1
        
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # 生成下载文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f'订单二维码打印表_{timestamp}.xlsx'
        
        return send_from_directory(
            os.path.dirname(temp_file.name), 
            os.path.basename(temp_file.name),
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Excel导出错误: {e}")
        # 如果标准导出失败，尝试简化版本
        try:
            return export_simple_excel()
        except Exception as e2:
            return jsonify({'error': f'导出Excel失败: {str(e)}'}), 500

def export_simple_excel():
    """简化版Excel导出（不包含图片）"""
    try:
        import pandas as pd
        
        # 获取所有订单数据
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT order_id, customer_name, order_date, amount, product_details
            FROM orders 
            ORDER BY order_date DESC
        ''')
        orders = cursor.fetchall()
        conn.close()
        
        if not orders:
            return jsonify({'error': '没有订单数据可导出'}), 404
        
        # 创建DataFrame
        df_data = []
        for order in orders:
            order_id, customer_name, order_date, amount, product_details = order
            df_data.append({
                '订单号': order_id,
                '客户姓名': customer_name,
                '订单日期': order_date,
                '金额': f'¥{amount:.2f}',
                '产品详情': product_details,
                '二维码文件': f'qrcodes/order_{order_id}.png',
                '查询链接': f'http://localhost:5000/order?order_id={order_id}'
            })
        
        df = pd.DataFrame(df_data)
        
        # 创建临时文件
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        df.to_excel(temp_file.name, index=False, engine='openpyxl')
        temp_file.close()
        
        # 生成下载文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f'订单列表_{timestamp}.xlsx'
        
        return send_from_directory(
            os.path.dirname(temp_file.name), 
            os.path.basename(temp_file.name),
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'简化Excel导出失败: {str(e)}'}), 500

@app.route('/api/profit_analysis')
@login_required
def get_profit_analysis():
    """获取盈亏分析数据（按时间和产品维度）"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        product_code = request.args.get('product_code')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取所有有成本记录的产品列表（使用子查询确保只获取最新成本记录的产品）
        cursor.execute('''
            WITH LatestCosts AS (
                SELECT 
                    product_code,
                    ROW_NUMBER() OVER (PARTITION BY product_code ORDER BY calculation_date DESC) as rn
                FROM production_costs
            )
            SELECT DISTINCT 
                lc.product_code,
                ii.item_name as product_name
            FROM LatestCosts lc
            LEFT JOIN inventory_items ii ON lc.product_code = ii.item_code
            WHERE lc.rn = 1
            ORDER BY lc.product_code
        ''')
        products = cursor.fetchall()
        
        # 构建订单查询
        query = '''
            WITH LatestCosts AS (
                SELECT 
                    product_code,
                    total_cost,
                    ROW_NUMBER() OVER (PARTITION BY product_code ORDER BY calculation_date DESC) as rn
                FROM production_costs
            )
            SELECT 
                o.order_id,
                o.product_code,
                ii.item_name as product_name,
                o.order_date,
                o.amount as sale_price,
                lc.total_cost,
                (o.amount - lc.total_cost) as profit,
                (o.amount - lc.total_cost) / o.amount * 100 as profit_rate
            FROM orders o
            LEFT JOIN LatestCosts lc ON o.product_code = lc.product_code AND lc.rn = 1
            LEFT JOIN inventory_items ii ON o.product_code = ii.item_code
            WHERE o.product_code IS NOT NULL
            AND lc.total_cost IS NOT NULL
        '''
        
        params = []
        if start_date:
            query += ' AND o.order_date >= ?'
            params.append(start_date)
        if end_date:
            query += ' AND o.order_date <= ?'
            params.append(end_date)
        if product_code:
            query += ' AND o.product_code = ?'
            params.append(product_code)
            
        query += ' ORDER BY o.order_date'
        
        cursor.execute(query, params)
        orders = cursor.fetchall()
        
        conn.close()
        
        # 处理数据
        orders_data = []
        for order in orders:
            orders_data.append({
                'order_id': order['order_id'],
                'product_code': order['product_code'],
                'product_name': order['product_name'] or order['product_code'],
                'order_date': order['order_date'],
                'sale_price': float(order['sale_price']),
                'total_cost': float(order['total_cost']),
                'profit': float(order['profit']),
                'profit_rate': float(order['profit_rate'])
            })
            
        # 按产品统计
        product_stats = {}
        for order in orders_data:
            product_key = order['product_code']
            if product_key not in product_stats:
                product_stats[product_key] = {
                    'product_code': product_key,
                    'product_name': order['product_name'],
                    'total_orders': 0,
                    'total_sales': 0,
                    'total_costs': 0,
                    'total_profit': 0,
                    'avg_profit_rate': 0
                }
            
            stats = product_stats[product_key]
            stats['total_orders'] += 1
            stats['total_sales'] += order['sale_price']
            stats['total_costs'] += order['total_cost']
            stats['total_profit'] += order['profit']
        
        # 计算平均利润率
        for stats in product_stats.values():
            if stats['total_sales'] > 0:
                stats['avg_profit_rate'] = (stats['total_profit'] / stats['total_sales']) * 100
        
        # 准备产品选择列表
        products_list = [{
            'product_code': p['product_code'],
            'product_name': p['product_name'] or p['product_code']
        } for p in products]
        
        return jsonify({
            'success': True,
            'orders': orders_data,
            'product_stats': list(product_stats.values()),
            'products': products_list
        })
        
    except Exception as e:
        print(f"❌ 获取盈亏分析数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/inventory_summary')
@login_required
def get_inventory_summary():
    """获取库存汇总信息"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取所有库存物料信息
        cursor.execute('''
            SELECT 
                item_code,
                item_name,
                item_category,
                unit,
                current_stock,
                weighted_avg_price,
                total_value,
                low_stock_threshold,
                warning_stock_threshold
            FROM inventory_items
            ORDER BY item_category, item_code
        ''')
        
        items = []
        total_items = 0
        total_stock = 0
        total_value = 0
        
        for row in cursor.fetchall():
            item = {
                'item_code': row['item_code'],
                'item_name': row['item_name'],
                'item_category': row['item_category'],
                'unit': row['unit'],
                'current_stock': float(row['current_stock']),
                'weighted_avg_price': float(row['weighted_avg_price']),
                'total_value': float(row['total_value']),
                'low_stock_threshold': int(row['low_stock_threshold']),
                'warning_stock_threshold': int(row['warning_stock_threshold'])
            }
            items.append(item)
            
            total_items += 1
            total_stock += item['current_stock']
            total_value += item['total_value']
        
        conn.close()
        
        return jsonify({
            'success': True,
            'items': items,
            'total_items': total_items,
            'total_stock': int(total_stock),
            'total_value': total_value
        })
        
    except Exception as e:
        print(f"❌ 获取库存汇总失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/cost_analysis')
@login_required
def get_cost_analysis():
    """获取成本分析报告API"""
    try:
        processor = OrderProcessor()
        result = processor.get_cost_analysis_report()
        
        if result['success']:
            return jsonify(result)
        else:
            return jsonify({'error': result['error']}), 500
        
    except Exception as e:
        return jsonify({'error': f'获取成本分析失败: {str(e)}'}), 500

@app.route('/api/cost_analysis_report')
@login_required
def get_cost_analysis_report():
    """获取成本分析报告数据"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取最新成本记录的产品列表
        cursor.execute('''
            WITH LatestProductCosts AS (
                SELECT 
                    product_code,
                    material_cost,
                    labor_cost,
                    management_cost,
                    transport_cost,
                    other_cost,
                    total_cost,
                    calculation_date,
                    ROW_NUMBER() OVER (PARTITION BY product_code ORDER BY calculation_date DESC) as rn
                FROM production_costs
            )
            SELECT 
                pc.*,
                ii.item_name
            FROM LatestProductCosts pc
            LEFT JOIN inventory_items ii ON pc.product_code = ii.item_code
            WHERE pc.rn = 1
            ORDER BY pc.product_code
        ''')
        
        products = cursor.fetchall()
        conn.close()
        
        if not products:
            return jsonify({
                'success': True,
                'data': [],
                'products': []
            })
        
        # 处理数据
        products_data = []
        for product in products:
            total_cost = float(product['total_cost'])
            products_data.append({
                'product_code': product['product_code'],
                'product_name': product['item_name'] or product['product_code'],
                'total_cost': total_cost,
                'cost_breakdown': [
                    {
                        'name': '材料成本',
                        'value': float(product['material_cost']),
                        'percentage': round(float(product['material_cost']) / total_cost * 100, 2) if total_cost > 0 else 0
                    },
                    {
                        'name': '人工成本',
                        'value': float(product['labor_cost']),
                        'percentage': round(float(product['labor_cost']) / total_cost * 100, 2) if total_cost > 0 else 0
                    },
                    {
                        'name': '管理成本',
                        'value': float(product['management_cost']),
                        'percentage': round(float(product['management_cost']) / total_cost * 100, 2) if total_cost > 0 else 0
                    },
                    {
                        'name': '运输成本',
                        'value': float(product['transport_cost']),
                        'percentage': round(float(product['transport_cost']) / total_cost * 100, 2) if total_cost > 0 else 0
                    },
                    {
                        'name': '其他成本',
                        'value': float(product['other_cost']),
                        'percentage': round(float(product['other_cost']) / total_cost * 100, 2) if total_cost > 0 else 0
                    }
                ]
            })
        
        # 准备产品选择列表
        products_list = [{
            'product_code': p['product_code'],
            'product_name': p['item_name'] or p['product_code']
        } for p in products]
        
        return jsonify({
            'success': True,
            'data': products_data,
            'products': products_list
        })
        
    except Exception as e:
        print(f"❌ 获取成本分析报告失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/calculate_product_cost', methods=['POST'])
@login_required
def calculate_product_cost():
    """计算产品成本API"""
    try:
        data = request.get_json()
        
        if not data or 'product_code' not in data:
            return jsonify({'success': False, 'error': '缺少产品编码参数'}), 400
            
        product_code = data['product_code']
        quantity = float(data.get('quantity', 1))
        labor_hours = float(data.get('labor_hours', 0))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取BOM清单
        cursor.execute('''
            SELECT 
                bi.material_code,
                bi.required_quantity as quantity,
                ii.weighted_avg_price,
                ii.item_name
            FROM bom_items bi 
            JOIN inventory_items ii ON bi.material_code = ii.item_code
            WHERE bi.product_code = ?
        ''', (product_code,))
        
        bom_items = cursor.fetchall()
        
        if not bom_items:
            return jsonify({'success': False, 'error': '找不到产品的BOM清单'}), 404
        
        # 计算材料成本
        material_cost = 0
        for item in bom_items:
            qty = float(item['quantity']) if item['quantity'] is not None else 0
            price = float(item['weighted_avg_price']) if item['weighted_avg_price'] is not None else 0
            material_cost += qty * price
        
        # 获取所有有效的成本配置项
        cursor.execute('''
            SELECT 
                item_name,
                item_type,
                default_value,
                unit,
                description
            FROM cost_config_items 
            WHERE is_active = 1
            ORDER BY id
        ''')
        configs = cursor.fetchall()
        
        # 初始化成本字典
        costs = {
            'material_cost': material_cost,  # 材料成本是基础成本
            'labor_cost': 0,                 # 人工成本将基于工时计算
            'management_cost': 0,            # 管理成本
            'transport_cost': 0,             # 运输成本
            'other_costs': {},               # 其他成本项
            'tax_rate': None,                # 税率配置
            'tax_cost': 0                    # 税费金额
        }
        
        # 处理每个成本配置项
        for config in configs:
            name = config['item_name']
            value = float(config['default_value'])
            cost_type = config['item_type']
            
            # 如果是税率配置，先保存起来，最后计算
            if name == '税费':
                costs['tax_rate'] = {
                    'name': name,
                    'value': value,
                    'type': cost_type
                }
                continue
            
            # 根据配置项名称和类型计算成本
            if name == '人工费率':
                if cost_type == 'fixed':
                    costs['labor_cost'] = labor_hours * value  # 固定值：每小时费率
                else:
                    costs['labor_cost'] = material_cost * (value / 100)  # 百分比：占材料成本比例
            elif name == '管理费率':
                if cost_type == 'fixed':
                    costs['management_cost'] = value  # 固定值
                else:
                    costs['management_cost'] = material_cost * (value / 100)  # 百分比
            elif name == '运输费率':
                if cost_type == 'fixed':
                    costs['transport_cost'] = value  # 固定值
                else:
                    costs['transport_cost'] = material_cost * (value / 100)  # 百分比
            else:  # 处理其他所有成本项（包括设备折旧费等）
                if cost_type == 'fixed':
                    costs['other_costs'][name] = value  # 固定值
                else:
                    costs['other_costs'][name] = material_cost * (value / 100)  # 百分比：基于材料成本
        
        # 计算中间总成本（不含税）
        subtotal_cost = (
            costs['material_cost'] + 
            costs['labor_cost'] + 
            costs['management_cost'] + 
            costs['transport_cost'] + 
            sum(costs['other_costs'].values())
        )
        
        # 计算税费（基于不含税总成本）
        if costs['tax_rate']:
            tax_config = costs['tax_rate']
            if tax_config['type'] == 'percentage':
                costs['tax_cost'] = subtotal_cost * (tax_config['value'] / 100)  # 基于不含税总成本计算
            else:
                costs['tax_cost'] = tax_config['value']  # 固定税费
        
        # 计算最终总成本（含税）
        total_cost = subtotal_cost + costs['tax_cost']
        unit_cost = total_cost / quantity if quantity > 0 else 0
        
        # 生成成本记录ID
        cost_id = f"COST_{product_code}"
        
        # 先删除该产品的旧成本记录
        cursor.execute('DELETE FROM production_costs WHERE product_code = ?', (product_code,))
        
        # 保存新的成本记录
        cursor.execute('''
            INSERT INTO production_costs (
                cost_id, 
                product_code, 
                material_cost,
                labor_cost,
                management_cost,
                transport_cost,
                other_cost,
                tax_cost,
                total_cost,
                quantity,
                unit_cost,
                calculation_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (
            cost_id,
            product_code,
            costs['material_cost'],
            costs['labor_cost'],
            costs['management_cost'],
            costs['transport_cost'],
            sum(costs['other_costs'].values()),  # 其他所有费用
            costs['tax_cost'],
            total_cost,
            quantity,
            unit_cost
        ))
        
        conn.commit()
        conn.close()
        
        # 准备成本明细
        cost_details = []
        
        # 1. 材料成本
        cost_details.append({
            'name': '材料成本',
            'value': costs['material_cost'],
            'type': 'fixed',
            'percentage': (costs['material_cost'] / total_cost * 100) if total_cost > 0 else 0
        })
        
        # 2. 人工成本
        labor_config = next((c for c in configs if c['item_name'] == '人工费率'), None)
        cost_details.append({
            'name': '人工成本',
            'value': costs['labor_cost'],
            'type': labor_config['item_type'] if labor_config else 'fixed',
            'percentage': (costs['labor_cost'] / total_cost * 100) if total_cost > 0 else 0
        })
        
        # 3. 管理成本
        management_config = next((c for c in configs if c['item_name'] == '管理费率'), None)
        cost_details.append({
            'name': '管理成本',
            'value': costs['management_cost'],
            'type': management_config['item_type'] if management_config else 'fixed',
            'percentage': (costs['management_cost'] / total_cost * 100) if total_cost > 0 else 0
        })
        
        # 4. 运输成本
        transport_config = next((c for c in configs if c['item_name'] == '运输费率'), None)
        cost_details.append({
            'name': '运输成本',
            'value': costs['transport_cost'],
            'type': transport_config['item_type'] if transport_config else 'fixed',
            'percentage': (costs['transport_cost'] / total_cost * 100) if total_cost > 0 else 0
        })
        
        # 5. 其他成本项
        for name, value in costs['other_costs'].items():
            config = next((c for c in configs if c['item_name'] == name), None)
            if config:
                cost_details.append({
                    'name': name,
                    'value': value,
                    'type': config['item_type'],
                    'percentage': (value / total_cost * 100) if total_cost > 0 else 0
                })
        
        # 6. 税费（如果有）
        if costs['tax_rate']:
            cost_details.append({
                'name': costs['tax_rate']['name'],
                'value': costs['tax_cost'],
                'type': costs['tax_rate']['type'],
                'percentage': (costs['tax_cost'] / total_cost * 100) if total_cost > 0 else 0,
                'is_tax': True
            })
        
        # 按照百分比降序排序
        cost_details.sort(key=lambda x: x['percentage'], reverse=True)
        
        return jsonify({
            'success': True,
            'cost': {
                'product_code': product_code,
                'total_cost': round(total_cost, 2),
                'subtotal_cost': round(subtotal_cost, 2),
                'unit_cost': round(unit_cost, 2),
                'quantity': quantity,
                'labor_hours': labor_hours,
                'cost_details': [{
                    **detail,
                    'value': round(detail['value'], 2),
                    'percentage': round(detail['percentage'], 2)
                } for detail in cost_details],
                'calculation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        })
        
    except Exception as e:
        print(f"❌ 计算产品成本失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/product_costs')
@login_required
def get_product_costs():
    """获取产品成本记录API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 使用子查询获取每个产品最新的成本记录
        cursor.execute('''
            WITH LatestCosts AS (
                SELECT 
                    product_code,
                    material_cost,
                    labor_cost,
                    management_cost,
                    transport_cost,
                    other_cost,
                    total_cost,
                    quantity,
                    unit_cost,
                    calculation_date,
                    ROW_NUMBER() OVER (PARTITION BY product_code ORDER BY calculation_date DESC) as rn
                FROM production_costs
            )
            SELECT 
                lc.*,
                ii.item_name as product_name
            FROM LatestCosts lc
            LEFT JOIN inventory_items ii ON lc.product_code = ii.item_code
            WHERE lc.rn = 1
            ORDER BY lc.calculation_date DESC
        ''')
        
        costs = cursor.fetchall()
        conn.close()
        
        costs_list = []
        for cost in costs:
            costs_list.append({
                'product_code': cost['product_code'],
                'product_name': cost['product_name'] or cost['product_code'],
                'material_cost': float(cost['material_cost']),
                'labor_cost': float(cost['labor_cost']),
                'management_cost': float(cost['management_cost']),
                'transport_cost': float(cost['transport_cost']),
                'other_cost': float(cost['other_cost']),
                'total_cost': float(cost['total_cost']),
                'quantity': float(cost['quantity']),
                'unit_cost': float(cost['unit_cost']),
                'calculation_date': cost['calculation_date']
            })
        
        return jsonify({
            'success': True,
            'costs': costs_list
        })
        
    except Exception as e:
        print(f"❌ 获取产品成本记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/update_cost_config', methods=['POST'])
@login_required
def update_cost_config():
    """更新成本配置API"""
    try:
        data = request.get_json()
        
        if not data or 'config_type' not in data or 'config_value' not in data:
            return jsonify({'error': '缺少配置参数'}), 400
        
        config_type = data['config_type']
        config_value = float(data['config_value'])
        
        processor = OrderProcessor()
        result = processor.update_cost_config(config_type, config_value)
        
        if result['success']:
            return jsonify(result)
        else:
            return jsonify({'error': result['error']}), 400
        
    except Exception as e:
        return jsonify({'error': f'更新成本配置失败: {str(e)}'}), 500

@app.route('/api/update_item_thresholds', methods=['POST'])
@login_required
def update_item_thresholds():
    """更新物料阈值API"""
    try:
        data = request.get_json()
        
        if not data or 'item_id' not in data or 'low_threshold' not in data or 'warning_threshold' not in data:
            return jsonify({'error': '缺少参数'}), 400
        
        item_id = int(data['item_id'])
        low_threshold = float(data['low_threshold'])
        warning_threshold = float(data['warning_threshold'])
        
        processor = OrderProcessor()
        result = processor.update_item_thresholds(item_id, low_threshold, warning_threshold)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': f'更新阈值失败: {str(e)}'}), 500

@app.route('/api/bom_list')
@login_required
def get_bom_list():
    """获取BOM列表API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                b.id,
                b.product_code,
                i1.item_name as product_name,
                b.material_code,
                i2.item_name as material_name,
                b.required_quantity,
                b.unit,
                b.notes,
                b.created_at
            FROM bom_items b
            LEFT JOIN inventory_items i1 ON b.product_code = i1.item_code
            LEFT JOIN inventory_items i2 ON b.material_code = i2.item_code
            ORDER BY b.product_code, b.material_code
        ''')
        
        bom_items = cursor.fetchall()
        conn.close()
        
        bom_list = []
        for item in bom_items:
            bom_list.append({
                'id': item['id'],
                'product_code': item['product_code'],
                'product_name': item['product_name'] or '未知产品',
                'material_code': item['material_code'],
                'material_name': item['material_name'] or '未知物料',
                'required_quantity': float(item['required_quantity']),
                'unit': item['unit'],
                'notes': item['notes'] or '',
                'created_at': item['created_at']
            })
        
        return jsonify({
            'success': True,
            'bom_list': bom_list,
            'count': len(bom_list)
        })
        
    except Exception as e:
        return jsonify({'error': f'获取BOM列表失败: {str(e)}'}), 500

@app.route('/api/bom_item', methods=['POST', 'PUT', 'DELETE'])
@login_required
def manage_bom_item():
    """管理BOM项目API - 增删改"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if request.method == 'POST':
            # 添加新的BOM项目
            required_fields = ['product_code', 'material_code', 'required_quantity']
            for field in required_fields:
                if field not in data:
                    return jsonify({'error': f'缺少必需字段: {field}'}), 400
            
            # 检查并自动注册产品
            product_code = data['product_code']
            cursor.execute('SELECT item_code FROM inventory_items WHERE item_code = ?', (product_code,))
            if not cursor.fetchone():
                # 插入新产品
                cursor.execute('''
                    INSERT INTO inventory_items (
                        item_code, 
                        item_name, 
                        item_category,
                        unit,
                        current_stock,
                        weighted_avg_price,
                        total_value,
                        low_stock_threshold,
                        warning_stock_threshold
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    product_code,
                    data.get('product_name', product_code),  # 使用编码作为默认名称
                    '产品',  # 明确标记为产品类别
                    '个',
                    0,  # 初始库存
                    0,  # 初始平均价格
                    0,  # 初始总价值
                    10, # 默认低库存阈值
                    20  # 默认警告阈值
                ))
                print(f"✅ 自动注册产品: {product_code}")
            
            # 检查并自动注册原料
            material_code = data['material_code']
            cursor.execute('SELECT item_code FROM inventory_items WHERE item_code = ?', (material_code,))
            if not cursor.fetchone():
                # 根据编码前缀判断物料类型
                material_category = '原材料'
                if material_code.startswith('PKG'):
                    material_category = '包装'
                elif material_code.startswith('PART'):
                    material_category = '配件'
                
                # 插入新原料
                cursor.execute('''
                    INSERT INTO inventory_items (
                        item_code, 
                        item_name, 
                        item_category,
                        unit,
                        current_stock,
                        weighted_avg_price,
                        total_value,
                        low_stock_threshold,
                        warning_stock_threshold
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    material_code,
                    data.get('material_name', material_code),  # 使用编码作为默认名称
                    material_category,
                    data.get('unit', '个'),
                    0,  # 初始库存
                    0,  # 初始平均价格
                    0,  # 初始总价值
                    100, # 默认低库存阈值
                    200  # 默认警告阈值
                ))
                print(f"✅ 自动注册原料: {material_code} ({material_category})")
            
            # 添加BOM项目
            cursor.execute('''
                INSERT INTO bom_items (product_code, material_code, required_quantity, unit, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                product_code,
                material_code,
                float(data['required_quantity']),
                data.get('unit', '个'),
                data.get('notes', '')
            ))
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'BOM项目添加成功'
            })
            
        elif request.method == 'PUT':
            # 更新BOM项目
            if 'id' not in data:
                return jsonify({'error': '缺少BOM项目ID'}), 400
            
            update_fields = []
            update_values = []
            
            if 'product_code' in data:
                update_fields.append('product_code = ?')
                update_values.append(data['product_code'])
            
            if 'material_code' in data:
                update_fields.append('material_code = ?')
                update_values.append(data['material_code'])
            
            if 'required_quantity' in data:
                update_fields.append('required_quantity = ?')
                update_values.append(float(data['required_quantity']))
            
            if 'unit' in data:
                update_fields.append('unit = ?')
                update_values.append(data['unit'])
            
            if 'notes' in data:
                update_fields.append('notes = ?')
                update_values.append(data['notes'])
            
            if not update_fields:
                return jsonify({'error': '没有要更新的字段'}), 400
            
            update_values.append(data['id'])
            
            cursor.execute(f'''
                UPDATE bom_items 
                SET {', '.join(update_fields)}
                WHERE id = ?
            ''', update_values)
            
            if cursor.rowcount == 0:
                return jsonify({'error': 'BOM项目不存在'}), 404
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'BOM项目更新成功'
            })
            
        elif request.method == 'DELETE':
            # 删除BOM项目
            if 'id' not in data:
                return jsonify({'error': '缺少BOM项目ID'}), 400
            
            cursor.execute('DELETE FROM bom_items WHERE id = ?', (data['id'],))
            
            if cursor.rowcount == 0:
                return jsonify({'error': 'BOM项目不存在'}), 404
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'BOM项目删除成功'
            })
        
    except Exception as e:
        return jsonify({'error': f'BOM项目操作失败: {str(e)}'}), 500

@app.route('/api/bom_item/<int:bom_id>')
@login_required
def get_bom_item(bom_id):
    """获取单个BOM项目API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                b.id,
                b.product_code,
                i1.item_name as product_name,
                b.material_code,
                i2.item_name as material_name,
                b.required_quantity,
                b.unit,
                b.notes,
                b.created_at
            FROM bom_items b
            LEFT JOIN inventory_items i1 ON b.product_code = i1.item_code
            LEFT JOIN inventory_items i2 ON b.material_code = i2.item_code
            WHERE b.id = ?
        ''', (bom_id,))
        
        bom_item = cursor.fetchone()
        conn.close()
        
        if not bom_item:
            return jsonify({'error': 'BOM项目不存在'}), 404
        
        item_data = {
            'id': bom_item['id'],
            'product_code': bom_item['product_code'],
            'product_name': bom_item['product_name'] or '未知产品',
            'material_code': bom_item['material_code'],
            'material_name': bom_item['material_name'] or '未知物料',
            'required_quantity': float(bom_item['required_quantity']),
            'unit': bom_item['unit'],
            'notes': bom_item['notes'] or '',
            'created_at': bom_item['created_at']
        }
        
        return jsonify({
            'success': True,
            'bom_item': item_data
        })
        
    except Exception as e:
        return jsonify({'error': f'获取BOM项目失败: {str(e)}'}), 500

@app.route('/api/inventory_items')
@login_required
def get_inventory_items():
    """获取库存物品详情API"""
    try:
        category = request.args.get('category')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if category:
            cursor.execute('''
                SELECT 
                    id,
                    item_code,
                    item_name,
                    item_category,
                    unit,
                    current_stock,
                    weighted_avg_price,
                    total_value,
                    low_stock_threshold,
                    warning_stock_threshold,
                    last_updated
                FROM inventory_items 
                WHERE item_category = ?
                ORDER BY item_name
            ''', (category,))
        else:
            cursor.execute('''
                SELECT 
                    id,
                    item_code,
                    item_name,
                    item_category,
                    unit,
                    current_stock,
                    weighted_avg_price,
                    total_value,
                    low_stock_threshold,
                    warning_stock_threshold,
                    last_updated
                FROM inventory_items 
                ORDER BY item_category, item_name
            ''')
        
        items = cursor.fetchall()
        conn.close()
        
        items_list = []
        for item in items:
            items_list.append({
                'id': item['id'],
                'item_code': item['item_code'],
                'item_name': item['item_name'],
                'item_category': item['item_category'],
                'unit': item['unit'],
                'current_stock': float(item['current_stock']),
                'weighted_avg_price': float(item['weighted_avg_price']),
                'total_value': float(item['total_value']),
                'low_stock_threshold': int(item['low_stock_threshold'] or 100),
                'warning_stock_threshold': int(item['warning_stock_threshold'] or 200),
                'last_updated': item['last_updated']
            })
        
        return jsonify({
            'success': True,
            'items': items_list,
            'count': len(items_list),
            'category': category
        })
        
    except Exception as e:
        return jsonify({'error': f'获取库存明细失败: {str(e)}'}), 500

@app.route('/api/materials')
@login_required
def get_materials_list():
    """获取所有原料列表API（用于BOM下拉选择）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                item_code,
                item_name,
                item_category,
                unit
            FROM inventory_items 
            WHERE item_category IN ('原材料', '包装', '配件')
            ORDER BY 
                CASE item_category 
                    WHEN '原材料' THEN 1
                    WHEN '包装' THEN 2
                    WHEN '配件' THEN 3
                    ELSE 4
                END,
                item_code
        ''')
        
        materials = cursor.fetchall()
        conn.close()
        
        materials_list = []
        for material in materials:
            materials_list.append({
                'code': material['item_code'],
                'name': material['item_name'],
                'category': material['item_category'],
                'unit': material['unit']
            })
        
        return jsonify({
            'success': True,
            'materials': materials_list,
            'count': len(materials_list)
        })
        
    except Exception as e:
        return jsonify({'error': f'获取原料列表失败: {str(e)}'}), 500

@app.route('/api/products_with_bom')
@login_required
def get_products_with_bom():
    """获取有BOM清单的产品列表API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 直接从BOM表中获取唯一的产品编码
        cursor.execute('''
            SELECT DISTINCT 
                b.product_code as code,
                i.item_name as name
            FROM bom_items b
            LEFT JOIN inventory_items i ON b.product_code = i.item_code
            ORDER BY b.product_code
        ''')
        
        products = cursor.fetchall()
        conn.close()
        
        products_list = [{
            'code': product['code'],
            'name': product['name'] if product['name'] else product['code']
        } for product in products]
        
        return jsonify({
            'success': True,
            'products': products_list,
            'count': len(products_list)
        })
        
    except Exception as e:
        print(f"❌ 获取产品列表失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order_profit_report')
@login_required
def get_order_profit_report():
    """获取订单盈亏报告数据"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        product_code = request.args.get('product_code')
        
        # 构建查询条件
        query = '''
            WITH LatestCosts AS (
                SELECT 
                    product_code,
                    total_cost,
                    ROW_NUMBER() OVER (PARTITION BY product_code ORDER BY calculation_date DESC) as rn
                FROM production_costs
            )
            SELECT 
                o.order_id,
                o.product_code,
                o.amount as sale_price,
                lc.total_cost,
                ii.item_name as product_name,
                o.order_date,
                (o.amount - lc.total_cost) as profit,
                CASE 
                    WHEN o.amount = 0 THEN 0 
                    ELSE ((o.amount - lc.total_cost) / o.amount * 100) 
                END as profit_rate
            FROM orders o
            LEFT JOIN LatestCosts lc ON o.product_code = lc.product_code AND lc.rn = 1
            LEFT JOIN inventory_items ii ON o.product_code = ii.item_code
            WHERE o.product_code IS NOT NULL
            AND lc.total_cost IS NOT NULL
        '''
        params = []
        
        if start_date:
            query += ' AND date(o.order_date) >= date(?)'
            params.append(start_date)
        if end_date:
            query += ' AND date(o.order_date) <= date(?)'
            params.append(end_date)
        if product_code:
            query += ' AND o.product_code = ?'
            params.append(product_code)
            
        query += ' ORDER BY o.order_date DESC'
        
        cursor.execute(query, params)
        orders = cursor.fetchall()
        conn.close()
        
        if not orders:
            return jsonify({
                'success': True,
                'orders': []
            })
        
        orders_list = []
        for order in orders:
            profit = float(order['sale_price']) - float(order['total_cost'])
            orders_list.append({
                'order_id': order['order_id'],
                'product_code': order['product_code'],
                'product_name': order['product_name'] or order['product_code'],
                'sale_price': float(order['sale_price']),
                'total_cost': float(order['total_cost']),
                'profit': profit,
                'profit_rate': float(order['profit_rate']),
                'order_date': order['order_date']
            })
        
        return jsonify({
            'success': True,
            'orders': orders_list
        })
        
    except Exception as e:
        print(f"❌ 获取订单盈亏报告失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def process_orders_excel(file_path):
    """处理订单Excel文件"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 读取Excel文件
        df = pd.read_excel(file_path)
        
        # 检查必要的列是否存在
        required_columns = ['订单号', '产品编码', '金额']
        for col in required_columns:
            if col not in df.columns:
                return {
                    'success': False,
                    'error': f'Excel文件缺少必要的列: {col}'
                }
        
        # 检查订单号是否有重复
        duplicate_orders = []
        for order_id in df['订单号']:
            cursor.execute('SELECT order_id FROM orders WHERE order_id = ?', (order_id,))
            if cursor.fetchone():
                duplicate_orders.append(order_id)
        
        if duplicate_orders:
            return {
                'success': False,
                'error': f'以下订单号已存在，不能重复导入: {", ".join(duplicate_orders)}'
            }
        
        # 处理数据并插入数据库
        for _, row in df.iterrows():
            order_id = str(row['订单号'])
            product_code = str(row['产品编码'])
            amount = float(row['金额'])
            
            cursor.execute('''
                INSERT INTO orders (order_id, product_code, amount, order_date)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ''', (order_id, product_code, amount))
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'message': f'成功导入 {len(df)} 条订单数据'
        }
        
    except Exception as e:
        print(f"❌ 处理订单Excel失败: {str(e)}")
        if conn:
            conn.close()
        return {
            'success': False,
            'error': f'处理订单Excel失败: {str(e)}'
        }

@app.route('/api/cost_config_items')
@login_required
def get_cost_config_items():
    """获取成本配置项API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                id,
                item_name,
                item_type,
                default_value,
                unit,
                description,
                is_active
            FROM cost_config_items 
            WHERE is_active = 1
            ORDER BY id
        ''')
        
        items = cursor.fetchall()
        conn.close()
        
        items_list = []
        for item in items:
            items_list.append({
                'id': item['id'],
                'name': item['item_name'],
                'type': item['item_type'],
                'default_value': float(item['default_value']),
                'unit': item['unit'],
                'description': item['description'] or ''
            })
        
        return jsonify({
            'success': True,
            'items': items_list
        })
        
    except Exception as e:
        print(f"❌ 获取成本配置项失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/cost_config_items', methods=['POST'])
@login_required
def add_cost_config_item():
    """添加成本配置项API"""
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO cost_config_items 
            (item_name, item_type, default_value, unit, description)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data['name'],
            data['type'],
            float(data['default_value']),
            data['unit'],
            data.get('description', '')
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '成本配置项添加成功'
        })
        
    except Exception as e:
        print(f"❌ 添加成本配置项失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/cost_config_items/<int:item_id>', methods=['PUT'])
@login_required
def update_cost_config_item(item_id):
    """更新成本配置项API"""
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE cost_config_items 
            SET item_name = ?, item_type = ?, default_value = ?, 
                unit = ?, description = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['name'],
            data['type'],
            float(data['default_value']),
            data['unit'],
            data.get('description', ''),
            item_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '成本配置项更新成功'
        })
        
    except Exception as e:
        print(f"❌ 更新成本配置项失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/cost_config_items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_cost_config_item(item_id):
    """删除成本配置项API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE cost_config_items 
            SET is_active = 0, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (item_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '成本配置项删除成功'
        })
        
    except Exception as e:
        print(f"❌ 删除成本配置项失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/cost_config_items/<int:item_id>', methods=['GET'])
@login_required
def get_cost_config_item(item_id):
    """获取单个成本配置项API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                id,
                item_name,
                item_type,
                default_value,
                unit,
                description
            FROM cost_config_items 
            WHERE id = ? AND is_active = 1
        ''', (item_id,))
        
        item = cursor.fetchone()
        conn.close()
        
        if not item:
            return jsonify({
                'success': False,
                'error': '找不到指定的成本配置项'
            }), 404
        
        return jsonify({
            'success': True,
            'item': {
                'id': item['id'],
                'name': item['item_name'],
                'type': item['item_type'],
                'default_value': float(item['default_value']),
                'unit': item['unit'],
                'description': item['description'] or ''
            }
        })
        
    except Exception as e:
        print(f"❌ 获取成本配置项失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.errorhandler(404)
def not_found(error):
    """404错误处理"""
    return jsonify({
        'error': '页面不存在',
        'message': '请求的资源未找到'
    }), 404

@app.errorhandler(500)
def internal_error(error):
    """500错误处理"""
    return jsonify({
        'error': '服务器内部错误',
        'message': '服务器遇到了意外情况'
    }), 500

def init_app():
    """初始化应用"""
    try:
        # 确保上传目录存在
        os.makedirs('uploads', exist_ok=True)
        os.makedirs('qrcodes', exist_ok=True)
        
        # 初始化数据库（如果不存在）
        if not os.path.exists('orders.db'):
            print("📦 首次运行，创建新的数据库...")
            # 初始化数据库
            processor = OrderProcessor()
            if processor.init_database():
                print("✅ 数据库表创建成功")
                # 初始化示例数据
                if processor.init_sample_data():
                    print("✅ 示例数据初始化成功")
                else:
                    print("❌ 示例数据初始化失败")
            else:
                print("❌ 数据库初始化失败")
        else:
            print("✅ 数据库已存在，跳过初始化")
        
        print("✅ 应用初始化完成")
    except Exception as e:
        print(f"❌ 应用初始化失败: {str(e)}")
        raise e

# 在应用启动时初始化
init_app()

if __name__ == '__main__':
    import socket
    import os
    
    # 获取本机IP地址
    def get_local_ip():
        try:
            # 连接到一个远程地址来获取本机IP
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
            return local_ip
        except:
            return "127.0.0.1"
    
    # 检查数据库文件是否存在
    if not os.path.exists(DB_FILE):
        print(f"警告: 数据库文件 {DB_FILE} 不存在！")
        print("请先运行 excel_processor.py 来创建数据库和处理数据")
    
    # 确保二维码目录存在
    if not os.path.exists(QR_DIR):
        os.makedirs(QR_DIR)
        print(f"创建二维码目录: {QR_DIR}")
    
    # 云部署环境检测
    port = int(os.environ.get('PORT', 5000))
    host = '0.0.0.0' if 'PORT' in os.environ else '0.0.0.0'
    debug = os.environ.get('FLASK_ENV', 'development') != 'production'
    
    if 'PORT' in os.environ:
        # 云部署环境
        print("🌐 检测到云部署环境，启动生产模式...")
        print(f"🚀 应用将在端口 {port} 启动")
        print("📱 部署完成后，您将获得一个公网访问URL")
    else:
        # 本地开发环境
        local_ip = get_local_ip()
        
        print("🚀 订单二维码查询系统启动中...")
        print("=" * 50)
        print(f"💻 电脑端访问: http://localhost:{port}")
        print(f"💻 电脑端访问: http://127.0.0.1:{port}")
        print(f"📱 手机端访问: http://{local_ip}:{port}")
        print(f"🌐 局域网访问: http://{local_ip}:{port}")
        print("=" * 50)
        print()
        print("📱 手机使用步骤:")
        print("1. 确保手机和电脑连接同一WiFi网络")
        print(f"2. 手机浏览器输入: {local_ip}:{port}")
        print("3. 或直接扫描生成的二维码")
        print()
        print("🔧 功能特性:")
        print("  📊 Excel文件上传处理")
        print("  🔗 自动生成订单二维码")
        print("  📱 手机扫码查询订单")
        print("  🔍 网页手动查询订单")
        print()
        print("📡 API端点:")
        print("  - GET  /order?order_id=XXX  查询订单")
        print("  - GET  /orders             查看所有订单")
        print("  - GET  /qrcode/<order_id>  获取二维码")
        print("  - POST /upload             上传Excel文件")
        print("  - GET  /health             健康检查")
        print()
        print(f"⚠️  防火墙提示：如果手机无法访问，请确保Windows防火墙允许Python访问网络")
        print()
    
    app.run(debug=debug, host=host, port=port) 

@app.route('/api/update_product_stock', methods=['POST'])
@login_required
def update_product_stock():
    """更新产品库存API"""
    try:
        data = request.get_json()
        
        if not data or 'product_code' not in data or 'quantity' not in data:
            return jsonify({'error': '缺少必要参数'}), 400
        
        product_code = data['product_code']
        quantity = int(data['quantity'])
        unit_price = float(data.get('unit_price')) if data.get('unit_price') is not None else None
        notes = data.get('notes')
        
        # 检查产品是否存在
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT item_code, item_name, item_category 
            FROM inventory_items 
            WHERE item_code = ? AND item_category = '产品'
        ''', (product_code,))
        
        product = cursor.fetchone()
        conn.close()
        
        if not product:
            return jsonify({'error': '产品不存在或不是产品类型'}), 404
        
        # 更新产品库存
        processor = OrderProcessor()
        success = processor.update_product_stock(
            product_code=product_code,
            quantity=quantity,
            unit_price=unit_price,
            notes=notes
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': f'产品 {product_code} 入库成功'
            })
        else:
            return jsonify({
                'success': False,
                'error': '入库失败，请检查系统日志'
            }), 500
            
    except Exception as e:
        print(f"❌ 更新产品库存失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/update_thresholds', methods=['POST'])
@login_required
def update_thresholds():
    """更新物料的库存阈值"""
    try:
        data = request.get_json()
        if not data or 'item_code' not in data:
            return jsonify({'error': '缺少物料编码'}), 400
            
        item_code = data['item_code']
        low_threshold = int(data.get('low_stock_threshold', 0))
        warning_threshold = int(data.get('warning_stock_threshold', 0))
        
        if warning_threshold < low_threshold:
            return jsonify({'error': '警告阈值必须大于低库存阈值'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 更新阈值
        cursor.execute('''
            UPDATE inventory_items 
            SET low_stock_threshold = ?,
                warning_stock_threshold = ?,
                last_updated = CURRENT_TIMESTAMP
            WHERE item_code = ?
        ''', (low_threshold, warning_threshold, item_code))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': '物料不存在'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '阈值设置已更新'
        })
        
    except Exception as e:
        print(f"❌ 更新库存阈值失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/update_inventory_item', methods=['POST'])
@login_required
def update_inventory_item():
    """更新物料基本信息"""
    try:
        data = request.get_json()
        if not data or 'item_code' not in data:
            return jsonify({'error': '缺少物料编码'}), 400
            
        item_code = data['item_code']
        item_name = data.get('item_name', '').strip()
        unit = data.get('unit', '').strip()
        category = data.get('item_category', '').strip()
        
        if not item_name or not unit or not category:
            return jsonify({'error': '物料名称、单位和分类不能为空'}), 400
        
        if category not in ['原材料', '包装', '配件', '产品']:
            return jsonify({'error': '无效的物料分类'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 更新物料信息
        cursor.execute('''
            UPDATE inventory_items 
            SET item_name = ?,
                unit = ?,
                item_category = ?,
                last_updated = CURRENT_TIMESTAMP
            WHERE item_code = ?
        ''', (item_name, unit, category, item_code))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': '物料不存在'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '物料信息已更新'
        })
        
    except Exception as e:
        print(f"❌ 更新物料信息失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500